#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
سكربت زحف يومي — نسخة 3: روابط تصنيف/براند حقيقية موثّقة بدل تخمين روابط بحث،
+ استخراج كل المنتجات بالصفحة (مو منتج واحد فقط) + تصدير Excel منسّق شامل.

لماذا التغيير: تأكدنا فعليًا أن أغلب هذي المتاجر مبنية على منصة سلة، وكروت
المنتجات فيها تُبنى بمكوّن جافاسكربت (salla-product-card) يختلف تفصيله
حسب ثيم كل متجر. بدل تخمين رابط بحث + سيلكتور دقيق (فشل بالتجربة السابقة)،
نستخدم روابط تصنيف/براند حقيقية تحققنا منها يدويًا، وننتظر تحميل الصفحة
كاملة (networkidle) ثم نفكك النص الظاهر بالكامل لاستخراج كل زوج
(اسم منتج + سعر) موجود، بدل الاعتماد على سيلكتور واحد هش.
"""

import json
import re
import time
import random
import datetime
from pathlib import Path
from urllib.parse import quote

import requests
from playwright.sync_api import sync_playwright

try:
    from pytrends.request import TrendReq
    HAS_TRENDS = True
except ImportError:
    HAS_TRENDS = False

# ============================================================
# 1) روابط المنافسين — روابط تصنيف/براند حقيقية تم التحقق منها فعليًا
#    (target_urls) بدل تخمين روابط بحث عامة. لو ما عندنا رابط مؤكد،
#    نستخدم رابط بحث كمحاولة أخيرة (fallback=True) بثقة أقل.
# ============================================================
COMPETITORS = [
    {"name": "mokab", "target_urls": ["https://mokab.com/ar/joda/brand-1898110310"], "verified": True},
    {"name": "ppowerstore", "target_urls": ["https://ppowerstore.com/categories/321177/joda"], "verified": True},
    {"name": "aletawiksa", "target_urls": ["https://aletawiksa.com/categories/1122279/joda-%D8%AC%D9%88%D8%AF%D8%A7"], "verified": True},
    {"name": "al3bor-telecom", "target_urls": ["https://al3bor-telecom.com/categories/1122335/%D8%B4%D8%B1%D9%83%D8%A9-%D8%AC%D9%88%D8%AF%D8%A7"], "verified": True},
    {"name": "iblackstores", "target_urls": ["https://iblackstores.com/ar/search?q=%D8%AC%D9%88%D8%AF%D8%A7"], "verified": False},
    {"name": "saada", "target_urls": ["https://saada.sa/ar/search?q=%D8%AC%D9%88%D8%AF%D8%A7"], "verified": False},
    {"name": "wjhtektelecome", "target_urls": ["https://wjhtektelecome.com/ar/search?q=%D8%AC%D9%88%D8%AF%D8%A7"], "verified": False},
    {"name": "hope", "target_urls": ["https://hope.sa/ar/search?q=%D8%AC%D9%88%D8%AF%D8%A7"], "verified": False},
    {"name": "alwan7", "target_urls": ["https://alwan7.com/search?q=%D8%AC%D9%88%D8%AF%D8%A7"], "verified": False},
    {"name": "sync", "target_urls": ["https://sync.sa/ar/search?q=%D8%AC%D9%88%D8%AF%D8%A7"], "verified": False},
    {"name": "noon", "target_urls": ["https://www.noon.com/saudi-ar/search/?q=%D8%AC%D9%88%D8%AF%D8%A7"], "verified": False},
    {"name": "jarir", "target_urls": ["https://www.jarir.com/sa-en/catalogsearch/result/?q=joda"], "verified": False},
    {"name": "extra", "target_urls": ["https://www.extra.com/ar-sa/search/?q=joda"], "verified": False},
]

AMAZON = {"name": "amazon_sa", "search_url": "https://www.amazon.sa/s?k={q}"}

# كل منتج نراقبه + الكلمات المفتاحية اللي لازم تظهر كلها بعنوان المنتج
# عشان نعتبره تطابق (يخلي المطابقة أدق من مجرد "يحتوي على")
WATCHLIST = [
    {"id": "joda_20000", "category": "بطاريات متنقلة", "label": "بطارية جودا 20000mAh", "must_include": ["جودا", "20000"]},
    {"id": "joda_harbo", "category": "منصات ومحطات شحن", "label": "منصة جودا هاربو لشحن البطاريات", "must_include": ["هاربو"]},
    {"id": "joda_voltlink", "category": "توصيلات كهربائية", "label": "توصيلة جودا فولت لينك", "must_include": ["فولت لينك"]},
    {"id": "joda_airvolt", "category": "منتجات السيارة", "label": "مضخة هواء جودا AirVolt / مضخة بدون بطارية", "must_include": ["مضخة"]},
    {"id": "ravpower_pb", "category": "بطاريات متنقلة", "label": "بطارية راف باور", "must_include": ["راف باور", "بانك"]},
    {"id": "joda_voltpad", "category": "شواحن جدارية", "label": "شاحن جودا فولت باد / VoltPad", "must_include": ["فولت باد"]},
    {"id": "joda_cable240", "category": "كيابل شحن", "label": "كيبل جودا 240 واط", "must_include": ["240"]},
    {"id": "joda_car_charger", "category": "شواحن سيارة", "label": "شاحن سيارة جودا", "must_include": ["شاحن سيارة"]},
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
}

OUTPUT_JSON = Path(__file__).resolve().parent.parent / "data" / "products.json"
OUTPUT_XLSX = Path(__file__).resolve().parent.parent / "data" / "تقرير_المنافسين.xlsx"
OUTPUT_RAW_JSON = Path(__file__).resolve().parent.parent / "data" / "raw_scraped_products.json"

PRICE_PATTERN = re.compile(r"(\d{1,4}(?:[.,]\d{1,2})?)\s*(?:ر\.?\s*س|SAR|ريال)", re.IGNORECASE)
ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")


def normalize_digits(text: str) -> str:
    return text.translate(ARABIC_DIGITS)


def extract_products_from_text(body_text: str):
    """
    يفكك النص الكامل الظاهر بصفحة (بعد التصيير الكامل بالمتصفح) إلى قائمة
    منتجات مرشّحة: كل سطر فيه نمط سعر (رقم + ر.س/ريال/SAR) يُعتبر 'نهاية بطاقة
    منتج'، والأسطر القليلة اللي قبله تُعتبر اسم المنتج المرشّح.
    """
    text = normalize_digits(body_text)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    products = []
    for i, line in enumerate(lines):
        m = PRICE_PATTERN.search(line)
        if not m:
            continue
        try:
            price = float(m.group(1).replace(",", "."))
        except ValueError:
            continue
        if price < 5 or price > 5000:  # استبعاد أرقام غير منطقية كسعر منتج
            continue
        # نجمع حتى 3 أسطر سابقة كمرشح لاسم المنتج (نتجنب أسطر قصيرة جدًا زي "٪" أو "0")
        title_parts = []
        for back in range(1, 4):
            idx = i - back
            if idx < 0:
                break
            candidate = lines[idx]
            if len(candidate) >= 8 and not PRICE_PATTERN.search(candidate):
                title_parts.insert(0, candidate)
            if len(title_parts) >= 1:
                break
        if title_parts:
            products.append({"title": title_parts[-1], "price": price})
    return products


def matches_watchlist(title: str, watch_item: dict) -> bool:
    t = title.lower()
    return all(term.lower() in t for term in watch_item["must_include"])


def scrape_page_playwright(page, url: str):
    try:
        page.goto(url, timeout=15000, wait_until="domcontentloaded")
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass
        page.wait_for_timeout(1500)
        body_text = page.locator("body").inner_text(timeout=5000)
        return body_text
    except Exception as e:
        print(f"  ⚠ فشل تحميل {url}: {e}")
        return None


def scrape_amazon(query: str):
    url = AMAZON["search_url"].format(q=requests.utils.quote(query))
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"  ⚠ فشل الاتصال بأمازون: {e}")
        return []
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(resp.text, "html.parser")
    items = soup.select("[data-component-type='s-search-result']")[:15]
    out = []
    for it in items:
        title_el = it.select_one("h2 span")
        price_el = it.select_one(".a-price .a-offscreen")
        if title_el and price_el:
            price_txt = "".join(c for c in price_el.get_text() if c.isdigit() or c == ".")
            try:
                price = float(price_txt)
            except ValueError:
                continue
            out.append({"title": title_el.get_text(strip=True), "price": price})
    return out


def get_trend_interest(keywords, geo="SA"):
    if not HAS_TRENDS:
        return {}
    results = {}
    for kw in keywords:
        success = False
        for attempt in range(2):
            try:
                pytrends = TrendReq(hl="ar-SA", tz=180, retries=2, backoff_factor=0.5)
                pytrends.build_payload([kw], timeframe="today 12-m", geo=geo)
                df = pytrends.interest_over_time()
                if not df.empty and kw in df.columns:
                    results[kw] = int(df[kw].mean())
                    success = True
                break
            except Exception as e:
                wait = random.uniform(5, 8) * (attempt + 1)
                print(f"  ⚠ Trends محاولة {attempt+1} فشلت لـ '{kw}' — إعادة محاولة بعد {wait:.0f} ثانية")
                time.sleep(wait)
        if not success:
            results[kw] = None
        time.sleep(random.uniform(2, 4))
    return results


def build_excel_report(watchlist_results, raw_products_by_competitor):
    """يبني ملف Excel منسّق شامل: ملخص + كل المنتجات المكتشفة لكل منافس."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    navy = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    normal_font = Font(name="Arial", size=10.5)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = PatternFill("solid", fgColor="E2EFDA")
    amber = PatternFill("solid", fgColor="FCE4D6")
    grey = PatternFill("solid", fgColor="F2F2F2")

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = navy
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
            cell.border = border

    # ---- شيت 1: الملخص ----
    ws1 = wb.active
    ws1.title = "الملخص"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells("A1:G1")
    ws1["A1"] = f"تقرير رصد المنافسين — تيك تايم — {datetime.date.today().isoformat()}"
    ws1["A1"].font = title_font
    ws1["A1"].fill = navy
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 26

    headers = ["المنتج المُراقَب", "الفئة", "عدد المنافسين الذين لديهم المنتج", "المنافسون", "أقل سعر", "متوسط السعر", "درجة اهتمام Trends"]
    for i, h in enumerate(headers, start=1):
        ws1.cell(row=3, column=i, value=h)
    style_header(ws1, 3, len(headers))

    r = 3
    for item in watchlist_results:
        r += 1
        vals = [
            item["name"], item["cat"], item["competitorsCount"],
            item["competitorsList"] or "—",
            f'{item["minPrice"]} ر.س' if item["minPrice"] else "—",
            f'{item["avgPrice"]} ر.س' if item["avgPrice"] else "—",
            item["trendInterest"] if item["trendInterest"] is not None else item["trendStatus"],
        ]
        for i, v in enumerate(vals, start=1):
            cell = ws1.cell(row=r, column=i, value=v)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top", readingOrder=2)
        fill = green if item["competitorsCount"] >= 2 else (amber if item["competitorsCount"] == 1 else grey)
        for i in range(1, len(headers) + 1):
            ws1.cell(row=r, column=i).fill = fill
        ws1.row_dimensions[r].height = 32

    widths1 = [30, 20, 16, 30, 14, 14, 18]
    for i, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ---- شيت 2: كل المنتجات المكتشفة (شفافية كاملة/تدقيق) ----
    ws2 = wb.create_sheet("كل المنتجات المكتشفة")
    ws2.sheet_view.rightToLeft = True
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "كل منتج تم رصده فعليًا عند كل منافس (للتدقيق والمراجعة)"
    ws2["A1"].font = title_font
    ws2["A1"].fill = navy
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    h2 = ["المنافس", "اسم المنتج (كما استُخرج)", "السعر", "موثوقية الرابط"]
    for i, h in enumerate(h2, start=1):
        ws2.cell(row=3, column=i, value=h)
    style_header(ws2, 3, len(h2))

    r = 3
    for comp_name, info in raw_products_by_competitor.items():
        for prod in info["products"]:
            r += 1
            vals = [comp_name, prod["title"], f'{prod["price"]} ر.س', "رابط مؤكد" if info["verified"] else "بحث تقديري"]
            for i, v in enumerate(vals, start=1):
                cell = ws2.cell(row=r, column=i, value=v)
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top", readingOrder=2)
            ws2.row_dimensions[r].height = 20

    widths2 = [18, 55, 12, 16]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_XLSX)


def main():
    print(f"🚀 بدء الزحف اليومي (نسخة 3 — روابط موثّقة) — {datetime.datetime.now().isoformat()}")

    all_keywords = list({w["label"] for w in WATCHLIST})
    trend_scores = get_trend_interest(all_keywords)

    raw_products_by_competitor = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent=HEADERS["User-Agent"], locale="ar-SA", viewport={"width": 1280, "height": 900})
        page = context.new_page()

        for comp in COMPETITORS:
            print(f"\n🌐 {comp['name']} ({'رابط موثّق' if comp['verified'] else 'بحث تقديري'})")
            all_found = []
            for url in comp["target_urls"]:
                body_text = scrape_page_playwright(page, url)
                if body_text:
                    found = extract_products_from_text(body_text)
                    all_found.extend(found)
                time.sleep(random.uniform(1.5, 3))
            print(f"  وجدنا {len(all_found)} منتج مرشّح بهذي الصفحة")
            raw_products_by_competitor[comp["name"]] = {"products": all_found, "verified": comp["verified"]}

        browser.close()

    # أمازون منفصل (يشتغل عبر requests العادي، أثبت نجاحه بالتشغيل السابق)
    amazon_products = []
    for w in WATCHLIST:
        amazon_products.extend(scrape_amazon(" ".join(w["must_include"])))
        time.sleep(random.uniform(1, 2))
    raw_products_by_competitor["amazon_sa"] = {"products": amazon_products, "verified": True}

    # مطابقة كل منتج مراقَب مع كل ما رُصد عند كل منافس
    watchlist_results = []
    for w in WATCHLIST:
        found_at, prices = [], []
        for comp_name, info in raw_products_by_competitor.items():
            comp_matches = [p for p in info["products"] if matches_watchlist(p["title"], w)]
            if comp_matches:
                found_at.append(comp_name)
                prices.extend(p["price"] for p in comp_matches)

        trend_val = trend_scores.get(w["label"])
        watchlist_results.append({
            "id": w["id"], "cat": w["category"], "name": w["label"],
            "competitorsCount": len(found_at),
            "competitorsList": ", ".join(found_at),
            "minPrice": round(min(prices), 2) if prices else None,
            "avgPrice": round(sum(prices) / len(prices), 2) if prices else None,
            "trendInterest": trend_val,
            "trendStatus": "متاح" if trend_val is not None else "تعذر الجلب (حظر مؤقت من جوجل، طبيعي أحيانًا)",
            "updated": datetime.date.today().isoformat(),
        })
        print(f"\n📊 {w['label']}: {len(found_at)} منافس — {found_at}")

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(watchlist_results, f, ensure_ascii=False, indent=2)

    with open(OUTPUT_RAW_JSON, "w", encoding="utf-8") as f:
        json.dump(raw_products_by_competitor, f, ensure_ascii=False, indent=2)

    build_excel_report(watchlist_results, raw_products_by_competitor)

    print(f"\n✅ تم الحفظ: {OUTPUT_JSON.name}, {OUTPUT_RAW_JSON.name}, {OUTPUT_XLSX.name}")


if __name__ == "__main__":
    main()
